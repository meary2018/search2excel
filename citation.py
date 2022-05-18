import requests
import json
import datetime
import pprint
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
#URLは手動で設定してね
# offsetはどこからスタートするのか，limit はどこまで表示するのかを意味してます
#paper/より後にある 謎の文字列は読む論文のid です
#citationsは被引用文献を指してます．
url='https://api.semanticscholar.org/graph/v1/paper/837f2b1c72eb3ae286a37efc165c111a724b533c/citations?offset=0&limit=13&fields=title,authors,abstract'
res=requests.get(url)
data=json.loads(res.text)
list_data=data['data']
# 先にエクセルファイル作ってね
wb = openpyxl.load_workbook('220518_paperdata.xlsx')
# シートも作ってね
sheet = wb['220518_2']
sheet.cell(row=1, column=1).value = "No."
sheet.cell(row=1, column=2).value = "Title"
sheet.cell(row=1, column=3).value = "Author"
sheet.cell(row=1, column=4).value = "Abstract"

l=1
k=0
for i in range(len(list_data)):
    author=""
    l+=1
    k+=1
    sheet.cell(row=l, column=1).value =l-1
    sheet.cell(row=l, column=2).value =list_data[i]['citingPaper']['title']
    author_data=list_data[i]['citingPaper']['authors']
    for j in range(len(author_data)):
        if j==(len(author_data)-1):
            author+=author_data[j]['name']
        else:
            author+=author_data[j]['name']+','
    sheet.cell(row=l,column=3).value=author
    sheet.cell(row=l,column=4).value=list_data[i]['citingPaper']['abstract']

# エクセルファイルの設定はここも忘れないように！
wb.save('./220518_paperdata.xlsx')
wb.close()
